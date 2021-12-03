from openpyxl import load_workbook
import requests
import io
import re
import pandas as pd
def to_bq_column(s):
    s1 = re.sub('[ \n\t-+]', '_', str(s))
    if s1.startswith('_'):
        s2 = 'c' + s2
    else:
        s2 = s1
    return s2

def _decompose(number):
    """Generate digits from `number` in base alphabet, least significants
    bits first.

    Since A is 1 rather than 0 in base alphabet, we are dealing with
    `number - 1` at each iteration to be able to extract the proper digits.
    """

    while number:
        number, remainder = divmod(number - 1, 26)
        yield remainder


def base_10_to_alphabet(number):
    """Convert a decimal number to its base alphabet representation"""

    return ''.join(
            chr(ord('A') + part)
            for part in _decompose(number)
    )[::-1]
def base_alphabet_to_10(letters):
    """Convert an alphabet number to its decimal representation"""

    return sum(
            (ord(letter) - ord('A') + 1) * 26**i
            for i, letter in enumerate(reversed(letters.upper()))
    )
def parse_dimensions(dimensions_string):
    l,t,r,b = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', dimensions_string).groups()
    return (base_alphabet_to_10(l), int(t), base_alphabet_to_10(r), int(b))
    

def worksheet_to_frame(sheet, columns):
    print(f'declared dimensions: {sheet.dimensions}')
    shape = parse_dimensions(sheet.dimensions)
    max_col = len(columns)        
    if max_col != shape[2]:
        print('warning: range does not match columns passed')
    max_row = min(shape[3], 10000)
    row_list = list(sheet.iter_rows(min_row=1, max_col=max_col, max_row=max_row, values_only=True))
    return pd.DataFrame(row_list, columns=columns)

def worksheets_to_unified_frame(book, columns, prefix='', debug=False):
    if columns is None:
        columns 
    fr = None
    # loop instead of one whole fr for memory reasons
    for sheet in book.worksheets:
        print(sheet.title)
        if sheet.title.startswith(prefix):
            n = worksheet_to_frame(sheet, columns)
            print(n.shape)
            if fr is None:
                fr = n
            else:
                fr = pd.concat([fr, n])
    return fr

def read(url):  

    print('Reading from network')
    resp = requests.get(url)
    resp.raise_for_status()
    content = resp.content
    return content

def normalize(s):
    import re
    return re.sub(r'[^a-zA-Z0-9]', '_', s)
def cached_read(url):  
    from datetime import datetime
    key = f"cache/{datetime.strftime(datetime.now(),'%Y%m%d')}_{normalize(url)}"
    print(key)
    try:
        with open(key, 'rb') as f:
            data = f.read()
        print('A', type(data))
        return data
    except:
        data = read(url)
        with open(key, 'wb') as f:
            f.write(data)
        print('B', type(data))
        return data


    print('Reading from network')
    resp = requests.get(url)
    resp.raise_for_status()
    content = resp.content
    return content

def load_workbook_from_url(url):
    content = cached_read(url)
    bio = io.BytesIO()
    bio.write(content)
    bio.seek(0)
    return load_workbook(bio)