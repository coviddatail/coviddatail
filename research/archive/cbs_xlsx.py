# -*- coding: utf-8 -*-
# ---
# jupyter:
#   jupytext:
#     formats: ipynb,py:light
#     text_representation:
#       extension: .py
#       format_name: light
#       format_version: '1.5'
#       jupytext_version: 1.11.2
#   kernelspec:
#     display_name: Python 3
#     language: python
#     name: python3
# ---

import requests
import pandas as pd
from openpyxl import load_workbook
import io


from ezchart import bokeh_xy

deaths_url = 'https://apis.cbs.gov.il/series/data/path'
params = {
    'id':'2,2,1,1,644',
    'format': 'json',
    'download': False,
    'startPeriod': '01-2005', 
    'endPeriod':  '12-2021',
    'time': 3
}

res = requests.get(deaths_url, params=params)
res.raise_for_status()
js = [res.json()]
c = 0
while c < js[0]['DataSet']['paging']['last_page']:
    c = c + 1
    res = requests.get(deaths_url, params=dict(page=c, **params))
    res.raise_for_status()
    js.append(res.json())

# + tags=[]
[idx for idx, x in enumerate(js) if x['DataSet'] is None]
# -

frs = []
for j in js:
    for idx, s in enumerate(j['DataSet']['Series']):
        fr = pd.DataFrame(s['obs'])
        for k, v in s.items():
            if k == 'obs':
                continue            
            fr[k] = f"{v.get('name', k)},{v.get('value', k)}" if isinstance(v, dict) else v
        fr['idx'] = idx                                  
        frs.append(fr)            

fr0 = pd.concat(frs)
fr0.TimePeriod = pd.to_datetime(fr0.TimePeriod)

fr0.groupby(['time']).size()

pd.concat(frs, axis=0).time.drop_duplicates()

frm = pd.concat(frsm, axis=0)
fry = pd.concat(frsy, axis=0)

frm2.TimePeriod.dt.month

frm2 = frm.sort_values(by='TimePeriod').drop_duplicates().sort_values(by='TimePeriod').set_index('TimePeriod')
fry2 = fry.sort_values(by='TimePeriod').drop_duplicates().sort_values(by='TimePeriod').set_index('TimePeriod')

fry3 = fry2.copy().sort_index()
fry4 = fry3.resample('M', ).mean().ffill()/12

bokeh_xy.Line(frm2.index, frm2.Value).datetime() + 
bokeh_xy.Line(fry4.index, fry4.Value, color ='red').datetime() 

from ezchart import bokeh_xy

import numpy as np

frm2.to_csv('cbs_deaths.csv')

# +
fr3 = fr2.copy()
fr3['Y'] = fr2.TimePeriod.dt.year
fr3['M'] = fr2.TimePeriod.dt.month

fr4 = fr3.groupby(['Y', 'M']).size().sort_values(by='Value').sort_values(by='Value').reset_index()
bokeh_xy.Dots(fr3.Y + fr3.M/100, fr3.Value)
# -

fr3.head(100)

bokeh_xy.Dots(fr2.Ti

frms

js[0]

frm.sort_values(by='TimePeriod')

frm

# !ls ~/Downloads/*.xlsx

cbs_1 = pd.read_excel('/Users/odedbadt/Downloads/p-1.xlsx')

# !pip install openpyxl

cbs_1.shape

# ### Excel

bio = io.BytesIO()

bio.write(content)

workbooks[3]


def worksheet_to_frame(ws):
    return pd.DataFrame([[x.value for x in y] for y in ws['A1:Z377']])


re.match('[A-Za-z0-9]*', 'אב')

import re
def keep_ascii(l):
    return [re.sub('[ \n\t-+]', '_', str(x)) for x in l if re.match('[A-Za-z0-9]+', str(x))]
def create_header(raw_frame):
    raw_header = raw_frame.iloc[8:12,:]
    raw_header = raw_header.loc[(~raw_header.isna()).sum(axis=1) > 1, :]
    raw_header = raw_header.T.ffill().T
    has_english = raw_header.apply(lambda s: s.str.match('[A-Za-z0-9]+') )
    raw_header = raw_header.loc[has_english.any(axis=1),:]    
    return raw_header.apply(lambda s: '_'.join(keep_ascii(s.drop_duplicates().dropna())))


def extract_data(raw_frame):
    vals = raw_frame.iloc[12:,:].copy()
    vals[0] = pd.to_datetime(vals[0])
    return vals.set_index(0)


def extract_frame(raw_frame):
    vals = extract_data(raw_frame).dropna()
    headers = create_header(raw_frame)
    vals.columns = headers[1:]
    return vals


def workbook_list_to_frames(workbook):
    frs = [extract_frame(worksheet_to_frame(w)) for w in workbook if w.title.startswith('20')]
    return frs


wb = workbooks[1][-4]


# + tags=[]
def excel_to_unified_frame(book):
    worksheets = book.worksheets
    frames = workbook_list_to_frames(worksheets)
    columns = frames[0].columns
    def set_headers(x):
        x.columns = columns
        return x
    return pd.concat([set_headers(x) for x in a]).sort_index().dropna()
workbooks = {
    
}
for j in range(1, 4):
    print(j)
    resp = requests.get(f'https://www.cbs.gov.il/he/publications/LochutTlushim/2020/p-{j}.xlsx')
    resp.raise_for_status()
    content = resp.content
    bio = io.BytesIO()
    bio.write(content)
    bio.seek(0)
    book = load_workbook(bio)
    workbooks[j] = excel_to_unified_frame(book)
    print(workbooks[j].index.max())


# + jupyter={"outputs_hidden": true} tags=[]
wb['A1:Z377']
# -

raw_frame.T



pd.concat([set_headers(x) for x in a]).sort_index().dropna()

# + tags=[]
a[1].index
# -

raw_frame.T.iloc[0:12,:]

len(c)

raw_frame.iloc[6,0]

','.join(raw_frame.iloc[0:10,1].dropna())

extract_frame(raw_frame.T.iloc[8:12])

c.value

.iloc[0:12,:]

book = load_workbook('/Users/odedbadt/Downloads/p-1.xlsx')

ws = book['2020']

years = [book[str(y)] for y in range(2000, 2022)]

year_frames = [pd.DataFrame([[x.value for x in y] for y in year['A:Z']]).T.iloc[12:,:] for year in years]

one_frame = pd.concat(year_frames)
header = pd.DataFrame([[x.value for x in y] for y in ws['A1:Z11']])
#titles = header.agg(lambda x: ','.join([str(y)for y in x if y]))
titles = ['t', 'total', 'total_males', 'total_females', 'jewish_total', 'jewish_males', 'jewish_females', 'total_arabics', 'male_arabics', 'female_arabics'] + header.iloc[10,:].values.tolist()[10:]
one_frame.columns = titles
one_frame['t'] = pd.to_datetime(one_frame.t)
one_frame.set_index('t', inplace=True)
one_frame.dropna(how='all', inplace=True)

one_frame = one_frame.loc[~(one_frame == 'z').any(axis=1),:]

one_frame.loc[(one_frame == 'z').any(axis=1),:]

one_frame = one_frame.astype(float)

weekly = one_frame.resample('w').sum()
weekly.to_csv('cbs_weekly_deaths.csv')

one_frame.to_csv('cbs_daily_deaths.csv')


